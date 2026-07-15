# -*- coding: utf-8 -*-
"""
rekap_resmi.py
--------------
Membuat SATU sheet Excel dengan format PERSIS seperti dokumen resmi
instansi: "REKAPITULASI ABSENSI DIGITAL KEJAKSAAN TINGGI JAWA TENGAH"
(acuan: file Rekap_Juni_2026__Bin_Datun_Pidmil_SP.xlsx, sheet "Pidmil").

Sesuai arahan pembimbing (revisi):
- Semua pegawai digabung dalam SATU sheet (tidak dipecah per Bidang).
- Kolom "Bidang" tetap ada di tabel, tetapi nilainya dikosongkan per
  pegawai (tidak diisi otomatis).
- Kolom "Cuti" menampilkan RINCIAN jenis cuti (mis. "CUTI ALASAN PENTING
  : 5 Hari, CUTI BESAR : 3 Hari"), bukan cuma angka total - diambil dari
  baris statistik tambahan yang ditemukan di PDF sumber (lihat
  extractor.py -> _adalah_baris_statistik / _bangun_ringkasan).
"""

from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

BULAN_INDO = {
    1: "JANUARI", 2: "FEBRUARI", 3: "MARET", 4: "APRIL", 5: "MEI", 6: "JUNI",
    7: "JULI", 8: "AGUSTUS", 9: "SEPTEMBER", 10: "OKTOBER", 11: "NOVEMBER", 12: "DESEMBER",
}

THIN = Side(style="thin", color="000000")
BORDER_ALL = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

FONT_TITLE = Font(name="Arial", size=11, bold=True, underline="single")
FONT_PERIODE = Font(name="Arial", size=11, bold=True)
FONT_INFO = Font(name="Calibri", size=11, bold=True)
FONT_HEADER = Font(name="Calibri", size=11, bold=True)
FONT_DATA = Font(name="Calibri", size=11, bold=False)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CENTER_NOWRAP = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

KOLOM_HEADER = ["No", "Nama", "Bidang", "NIP", "Terlambat & Pulang Cepat",
                "Tidak Absen Pulang", "Cuti", "Sakit", "TAK", "Ket"]
LEBAR_KOLOM = [5.67, 44.0, 7.89, 22.44, 10.55, 8.66, 22.0, 8.55, 7.55, 12.0]

NAMA_SHEET = "Rekapitulasi Absensi"


def _format_periode(tanggal_list):
    """list string 'dd/mm/yyyy' -> teks 'DD BULAN YYYY' (awal & akhir)."""
    if not tanggal_list:
        return "-", "-"
    tanggal_dt = []
    for t in tanggal_list:
        try:
            tanggal_dt.append(datetime.strptime(t, "%d/%m/%Y"))
        except (ValueError, TypeError):
            continue
    if not tanggal_dt:
        return "-", "-"
    awal, akhir = min(tanggal_dt), max(tanggal_dt)

    def fmt(d):
        return f"{d.day} {BULAN_INDO[d.month]} {d.year}"

    return fmt(awal), fmt(akhir)


def tulis_sheet_rekap_resmi(workbook, ringkasan_list, semua_tanggal, nama_bidang=""):
    """
    workbook       : objek openpyxl Workbook (dari pd.ExcelWriter(...).book)
    ringkasan_list : list dict hasil ekstraksi (lihat extractor.py -> _bangun_ringkasan)
    semua_tanggal  : list string tanggal 'dd/mm/yyyy' dari seluruh baris harian
                     (dipakai untuk menentukan periode awal/akhir otomatis)
    nama_bidang    : opsional. Jika diisi (mis. "PIDMIL"), SEMUA baris pegawai
                     di batch ini akan diberi nilai Bidang tersebut, dan nama
                     sheet memakai nama Bidang itu (mis. sheet "Pidmil").
                     Jika dikosongkan, kolom Bidang dibiarkan kosong seperti
                     sebelumnya (dipakai saat batch berisi campuran beberapa
                     Bidang berbeda sekaligus).
    """
    if not ringkasan_list:
        return

    nama_bidang = (nama_bidang or "").strip()

    periode_awal, periode_akhir = _format_periode(semua_tanggal)
    jumlah_hari_kerja_list = [r.get("Total Hari Kerja") for r in ringkasan_list if isinstance(r.get("Total Hari Kerja"), int)]
    jumlah_hari_kerja = max(jumlah_hari_kerja_list) if jumlah_hari_kerja_list else "-"

    nama_sheet = nama_bidang.title()[:31] if nama_bidang else NAMA_SHEET
    if nama_sheet in workbook.sheetnames:
        nama_sheet = f"{NAMA_SHEET} 2"
    ws = workbook.create_sheet(nama_sheet)

    for idx, lebar in enumerate(LEBAR_KOLOM, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = lebar

    # baris 1: judul
    ws.merge_cells("A1:J1")
    c = ws.cell(row=1, column=1, value="REKAPITULASI ABSENSI DIGITAL KEJAKSAAN TINGGI JAWA TENGAH")
    c.font = FONT_TITLE
    c.alignment = ALIGN_CENTER_NOWRAP
    ws.row_dimensions[1].height = 14.4

    # baris 2: periode
    ws.merge_cells("A2:J2")
    c = ws.cell(row=2, column=1, value=f"PERIODE : {periode_awal} s.d. {periode_akhir}")
    c.font = FONT_PERIODE
    c.alignment = ALIGN_CENTER_NOWRAP
    ws.row_dimensions[2].height = 14.4

    # baris 3: jumlah hari kerja
    ws.merge_cells("G3:J3")
    for col in range(1, 11):
        ws.cell(row=3, column=col).font = FONT_INFO
        ws.cell(row=3, column=col).alignment = ALIGN_CENTER
    c = ws.cell(row=3, column=7, value=f"JUMLAH HARI KERJA : {jumlah_hari_kerja} HARI")
    c.alignment = ALIGN_RIGHT
    ws.row_dimensions[3].height = 14.4

    # baris 4: header tabel
    for col, judul in enumerate(KOLOM_HEADER, start=1):
        c = ws.cell(row=4, column=col, value=judul)
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    ws.row_dimensions[4].height = 43.2

    # urutkan berdasarkan nama, digabung semua pegawai dalam satu tabel
    daftar_urut = sorted(ringkasan_list, key=lambda r: r.get("Nama", ""))

    for i, r in enumerate(daftar_urut, start=1):
        baris = 4 + i
        terlambat = r.get("Terlambat (Hari)") or 0
        pulang_cepat = r.get("Pulang Cepat (Hari)") or 0
        terlambat_pulang_cepat = terlambat + pulang_cepat
        tidak_absen_pulang = r.get("Tidak Absen Pulang (Hari)") or 0
        rincian_cuti = r.get("Rincian Cuti") or None
        sakit = r.get("Sakit (Hari)") or 0
        tak = r.get("Alpha (Hari)") or 0  # TAK diasumsikan setara dengan Alpha

        nilai_baris = [
            i,
            r.get("Nama", "-"),
            nama_bidang.upper() if nama_bidang else None,  # Bidang: diisi jika nama_bidang diberikan
            r.get("NIP", "-"),
            terlambat_pulang_cepat or None,
            tidak_absen_pulang or None,
            rincian_cuti,
            sakit or None,
            tak or None,
            None,  # Ket dikosongkan, sesuai contoh acuan
        ]
        for col, val in enumerate(nilai_baris, start=1):
            c = ws.cell(row=baris, column=col, value=val)
            c.font = FONT_DATA
            c.border = BORDER_ALL
            if col == 2:
                c.alignment = ALIGN_LEFT
            elif col == 7:
                c.alignment = ALIGN_LEFT_WRAP
            else:
                c.alignment = ALIGN_CENTER_NOWRAP
        ws.row_dimensions[baris].height = 22.5

    ws.freeze_panes = "A5"
